from openpyxl import load_workbook
import os
import re
from scipy.interpolate import CubicSpline, UnivariateSpline, PchipInterpolator
import matplotlib.pyplot as plt
import numpy as np
from openpyxl import Workbook
from scipy.optimize import curve_fit
from scipy.optimize import leastsq
# 创建交互式界面
from matplotlib.widgets import Slider
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import tkinter as tk
from tkinter import ttk

# 插值函数参数
file_path ='data//施工图-3Y-路线逐桩坐标表.xlsx'
method = 'cubic'
Univariate_s = 0.05
isDrawRadius = False

# 全局插值函数
def interpolate(x, y):
    x = np.array(x)
    y = np.array(y)

    # 检查x是否是严格递减的，如果是则翻转x和y
    if np.all(np.diff(x) < 0):
        x = x[::-1]
        y = y[::-1]

    # 对x和y进行排序
    sorted_indices = np.argsort(x)
    x_sorted = x[sorted_indices]
    y_sorted = y[sorted_indices]
    if method == 'cubic':
        f = CubicSpline(x_sorted, y_sorted)
    elif method == 'univariate':
        f = UnivariateSpline(x_sorted, y_sorted, s=Univariate_s)
    elif method == 'pchip':
        f = PchipInterpolator(x_sorted, y_sorted)
    else:
        raise ValueError("Invalid method specified")
    
    return f

def normalize_data(data):
    """
    归一化数据
    
    参数:
    - data: 输入的二维数组
    
    返回:
    - 归一化后的数据
    """
    x = data[:, 2].astype(float)
    y = data[:, 3].astype(float)

    x_min = min(x)
    x_max = max(x)
    y_min = min(y)
    y_max = max(y)
    
    x_normalized = (x - x_min) / (x_max - x_min)
    y_normalized = (y - y_min) / (y_max - y_min)
    
    normalized_data = np.vstack((x_normalized, y_normalized)).T
    
    return normalized_data


def extract_number(text):
    number_match = re.search(r'(\d+)\+(\d+\.?\d*)', text)
    if number_match:
        number_part1 = number_match.group(1)
        number_part2 = number_match.group(2)
        combined_number = number_part1 + number_part2
        return float(combined_number)
    else:
        return None


def extract_data_from_excel(file_path):
    # 将相对路径转换为绝对路径
    abs_file_path = os.path.abspath(file_path)
    
    # 加载 Excel 文件
    wb = load_workbook(abs_file_path)
    ws = wb.active

    # 用于存储结果的列表
    data = []

    # 遍历工作表中的每一行
    for row in ws.iter_rows(min_row=1, max_col=ws.max_column, max_row=ws.max_row):
        for cell in row:
            # 检查单元格的值是否包含"k"
            if isinstance(cell.value, str) and "k" in cell.value.lower():
                # 提取当前单元格及其右侧两个单元格的值
                stake = cell.value
                stake_number = extract_number(stake)  # 匹配整个单词
    
                x_value = ws.cell(row=cell.row, column=cell.column + 1).value
                y_value = ws.cell(row=cell.row, column=cell.column + 2).value
                if x_value is not None and y_value is not None:
                    # 将数据添加到结果列表中
                    data.append([stake, stake_number, float(x_value), float(y_value)])
     # 关闭 Excel 文件
    wb.close()

    # 根据 stake_number 进行排序
    sorted_data = sorted(data, key=lambda x: x[1])
    
    return np.asarray(sorted_data) 

def segment_data(data):
    segments = []
    current_segment = []

    # 初始方向设置为None
    direction = None

    for i in range(len(data)):
        if i > 0:
            current_value = data[i][2]
            previous_value = data[i-1][2]
            if direction is None:
                # 初始方向
                if current_value > previous_value:
                    direction = "increasing"
                elif current_value < previous_value:
                    direction = "decreasing"
            else:
                # 检查当前值与前一个值的关系
                if (direction == "increasing" and current_value < previous_value) or \
                   (direction == "decreasing" and current_value > previous_value):
                    # 方向变化，保存当前段落并开始新的段落
                    segments.append(current_segment)
                    current_segment = [data[i-1]]  # 包含段落变化点
                    # 更新方向
                    if current_value > previous_value:
                        direction = "increasing"
                    elif current_value < previous_value:
                        direction = "decreasing"

        current_segment.append(data[i])

    # 添加最后一个段落
    if current_segment:
        segments.append(current_segment)

    return segments


def find_max_curvature_indices(segments, x_values, y_values):
    max_curvature_indices = []
    for segment in segments:
        start_index = segment[0]
        end_index = segment[1]
        if  end_index - start_index < 3:
            continue
        segment_x = x_values[start_index:end_index+1]
        segment_y = y_values[start_index:end_index+1]
        segment_curvature = calculate_curvature(segment_x, segment_y)
        max_index = start_index + np.argmax(segment_curvature)
        max_curvature_indices.append(max_index)
    return max_curvature_indices


def circle_equation(x, center_x, center_y, radius):
    return np.sqrt(radius**2 - (x - center_x)**2) + center_y

def find_circle(max_curvature_indices, non_zero_segments, n, x_values, y_values):
    for i, segment in enumerate(non_zero_segments):
        start_index = segment[0]
        end_index = segment[1]
        if end_index - start_index < n:
            continue
        max_index = max_curvature_indices[i]
        start_index = max_index - n
        end_index = max_index + n
        if start_index < segment[0] or end_index >= segment[1]:
            continue
        segment_x = x_values[start_index:end_index+1]
        segment_y = y_values[start_index:end_index+1]

        # 拟合圆曲线
        try:
            popt, pcov = curve_fit(circle_equation, segment_x, segment_y, p0=[0, 0, 1])
            center_x, center_y, radius = popt
            # 计算拟合误差
            fitted_curve = circle_equation(segment_x, center_x, center_y, radius)
            error = np.mean(np.abs(fitted_curve - segment_y))
            print(f"Segment {i+1}: Center ({center_x}, {center_y}), Radius {radius}, Fitting Error {error}")
        except Exception as e:
            print(f"Segment {i+1}: Fitting Failed - {e}")

# 定义一个函数，用于拟合圆
def calc_R(xc, yc, x, y):
    return np.sqrt((x - xc)**2 + (y - yc)**2)

def f_2(c, x, y):
    Ri = calc_R(c[0], c[1], x, y)
    return Ri - Ri.mean()

def Spline_fit_byXY(length, x, y, d=1, show=True):
    """
    拟合斜率曲线
    
    参数:
    - file_path: LAS文件路径
    - method: 插值方法，可选值为 'cubic', 'univariate', 'pchip'，默认为 'cubic'
    
    返回:
    - 无，直接绘制插值曲线
    """
    # 归一化 x 和 y 坐标
    x_min = np.min(x)
    x_max = np.max(x)
    y_min = np.min(y)
    y_max = np.max(y)
    
    x_scaled = (x - x_min) / (x_max - x_min)
    y_scaled = (y - y_min) / (y_max - y_min)
    
    # 根据选择的方法进行样条插值
    f = interpolate(x_scaled, y_scaled)
    y_prime, y_second, y_third = compute_derivatives(f, x_scaled)
    zero_segments, non_zero_segments = find_zero_second_derivatives(x_scaled, y_second, 1e-1)
    #max_curvature_indices = find_max_curvature_indices(non_zero_segments, x_scaled, y_scaled)
    initial_params = []  # 初始参数值
    param_names = []  # 参数名称
    for i, segment in enumerate(non_zero_segments):
        initial_params.append(100)  # 根据实际情况设置初始值
        param_names.append(f'第{i+1}段圆曲线拟合点数')

    #n = 5  # 延伸的点数
#
    #find_circle(max_curvature_indices, non_zero_segments, n, x_scaled, y_scaled)
    ## 生成插值后的 x 值
    #num_points = int(length / d) + 1
    #x_interp_scaled = np.linspace(x_min, x_max, num_points)
    #
    ## 反归一化得到原始坐标
    #x_interp = x_interp_scaled
    #y_interp_scaled = f((x_interp_scaled - x_min) / (x_max - x_min))
    #y_interp = y_interp_scaled * (y_max - y_min) + y_min

    #save_segments_to_excel(zero_segments, result, 'zero_segments.xlsx')
    #save_segments_to_excel(non_zero_segments, result, 'non_zero_segments.xlsx')
    # 绘制坐标点图和拟合曲线图
    if show:
        #plot_spline_and_derivatives(x_scaled, y_scaled, f, method) # 绘制样条曲线及其一、二、三阶导数
        plot_spline_fit_segments(zero_segments, non_zero_segments, f, x_scaled, y_scaled, initial_params, param_names) # 绘制样条拟合的各段和拟合圆
    return zero_segments,non_zero_segments



def plot_spline_and_derivatives(x_scaled, y_scaled, f, method):
    """
    绘制样条曲线及其一、二、三阶导数
    
    参数:
    - x_scaled: 归一化的 x 坐标
    - y_scaled: 归一化的 y 坐标
    - f: 插值函数
    - method: 插值方法的名称
    
    返回:
    - 无，直接绘制图形
    """
    plt.figure()
    
    # 绘制原始数据点
    plt.plot(x_scaled, y_scaled, 'o', label='data points')
    
    # 绘制整条样条曲线
    x_fine = np.linspace(np.min(x_scaled), np.max(x_scaled), 500)
    y_fine = f(x_fine)
    plt.plot(x_fine, y_fine, '-', label=f'{method} spline fit')

    # 计算一阶导数
    f_prime = f.derivative()
    y_prime = f_prime(x_fine)
    plt.plot(x_fine, y_prime, '--', label=f'{method} spline first derivative')
    
    # 计算二阶导数
    f_second = f_prime.derivative()
    y_second = f_second(x_fine)
    plt.plot(x_fine, y_second, '-.', label=f'{method} spline second derivative')

    # 计算三阶导数
    f_third = f_second.derivative()
    y_third = f_third(x_fine)
    plt.plot(x_fine, y_third, ':', label=f'{method} spline third derivative')
    
    plt.legend()
    plt.xlabel('x (scaled)')
    plt.ylabel('y (scaled)')
    plt.title(f'{method} Spline Fit and Derivatives')
    plt.show()


def plot_spline_fit_segments(zero_segments, non_zero_segments, f, x_scaled, y_scaled, params, param_names):
    """
    绘制样条拟合的各段和拟合圆
    
    参数:
    - zero_segments: 零二阶导数的段的列表
    - non_zero_segments: 非零二阶导数的段的列表
    - f: 插值函数
    - x_scaled: 归一化的 x 坐标
    - y_scaled: 归一化的 y 坐标
    
    返回:
    - 无，直接绘制图形
    """
    global textbox1, textbox2, combobox  # 声明为全局变量
    # 创建Tkinter主窗口
    root = tk.Tk()
    root.title("Spline Fit and Derivatives")

    # 创建组合框
    #param_names = [f'Param {i+1}' for i in range(len(non_zero_segments))]
    #combobox = ttk.Combobox(root, values=param_names)
    #combobox.grid(row=0, column=0, padx=10, pady=10)
    #combobox.current(0)  # 默认选择第一个参数

    # 创建Matplotlib图表并嵌入到Tkinter窗口
    fig, ax = plt.subplots()
    plt.subplots_adjust(left=0.25, bottom=0.35)
    canvas = FigureCanvasTkAgg(fig, master=root)
    canvas_widget = canvas.get_tk_widget()
    canvas_widget.grid(row=0, column=0, columnspan=6, sticky=tk.NSEW)  # 使用 sticky 参数并跨越所有列

    # 让组件随着窗口大小变化而调整大小
    root.grid_rowconfigure(0, weight=1)
    root.grid_columnconfigure(0, weight=1)

    # 绘制初始图形
    # 根据 zero_segments 绘制不同颜色的样条曲线
    x_segments, y_segments = zero_segments_calculator(x_scaled, f, zero_segments)
    for x_segment, y_segment in zip(x_segments, y_segments):
        ax.plot(x_segment, y_segment, color='red')
    # 其他段的样条曲线颜色
    for i, segment in enumerate(non_zero_segments):
        start_index, end_index = segment
        x_segment = x_scaled[start_index:end_index+1]
        y_segment = f(x_segment)
        ax.plot(x_segment, y_segment, color='blue')
    initial_params = [100] * len(non_zero_segments)
    x_fits, y_fits, errors, centers = non_zero_segments_calculator(x_scaled, f, non_zero_segments, initial_params, None)
    bowsSum=[]
    for x_fit, y_fit in zip(x_fits, y_fits):
        bows = bow_calculator(x_fit, y_fit, 10)
        bowsSum.append(bows)
        if x_fit is not None and y_fit is not None:
            ax.plot(x_fit, y_fit, 'g-', linewidth=3)
            pass
    plt.legend()
    plt.xlabel('x (scaled)')
    plt.ylabel('y (scaled)')
    plt.title('Spline Fit and Derivatives')
    # 更新图形的回调函数
    def update(event=None):
        # 收集所有的textbox1和textbox2的值
        textbox1_values = []
        textbox2_values = []
        for control in controls:
            label, textbox = control
            # 判断当前文本框是textbox1还是textbox2
            if "中心点" in label.cget("text"):  # 假设标签文本以“第”开头
                value = int(textbox.get())  # 获取当前文本框的值并转换为浮点数
                textbox1_values.append(value)
            elif "长度" in label.cget("text"):  # 假设标签文本以“长度”开头
                value = int(textbox.get())  # 获取当前文本框的值并转换为浮点数
                textbox2_values.append(value)

        # 从文本框获取参数值
        new_params1 = textbox1_values  # 使用所有textbox1的值作为参数值
        new_params2 = textbox2_values  # 使用所有textbox2的值作为参数值
        # 从下拉框获取选项值
        selected_option = combobox.get()
        ax.clear()  # 清除之前的绘图
        #params = [slider.val for slider in sliders]
        x_segments, y_segments = zero_segments_calculator(x_scaled, f, zero_segments)
        x_fits, y_fits, errors, centers = non_zero_segments_calculator(x_scaled, f, non_zero_segments, new_params2, new_params1)
        
        # 重新绘图
        for x_segment, y_segment in zip(x_segments, y_segments):
            ax.plot(x_segment, y_segment, color='red')
        # 其他段的样条曲线颜色
        for i, segment in enumerate(non_zero_segments):
            start_index, end_index = segment
            x_segment = x_scaled[start_index:end_index+1]
            y_segment = f(x_segment)
            ax.plot(x_segment, y_segment, color='blue')
        for x_fit, y_fit in zip(x_fits, y_fits):
            if x_fit is not None and y_fit is not None:
                ax.plot(x_fit, y_fit, 'g-', linewidth=3)
                pass
        plt.legend()
        plt.xlabel('x (scaled)')
        plt.ylabel('y (scaled)')
        plt.title('Spline Fit and Derivatives')
        fig.canvas.draw_idle()  # 更新图形
    # 添加滑动条
    axcolor = 'lightgoldenrodyellow'
    sliders = []
    controls = []
    for i, param in enumerate(params):
        #ax_param = plt.axes([0.36, 0.01 + i*0.05, 0.3, 0.03], facecolor=axcolor)
        #slider = Slider(ax_param, f'Param {param_names[i]}', param-1000, param+1000, valinit=param, valstep=1)
        #sliders.append(slider)

        label1 = tk.Label(root, text=f'第{i}段圆曲线中心点位置:')
        label1.grid(row=i+1, column=0, padx=(10, 10), pady=(5, 0), sticky='e')  # 调整标签的位置和间距
        # 创建文本框
        textbox1 = tk.Entry(root)
        textbox1.insert(tk.END, str(centers[i]))  # 设置默认值
        textbox1.grid(row=i+1, column=1, padx=(10, 10), pady=(5, 0))  # 调整文本框的位置和间距
        controls.append((label1, textbox1))
        textbox1.focus_set()  # 设置焦点到textbox1
        # 创建标签
        label2 = tk.Label(root, text=f'第{i}段圆曲线长度:')
        label2.grid(row=i+1, column=2, padx=(10, 10), pady=(5, 0), sticky='e')  # 调整标签的位置和间距
        # 创建文本框
        textbox2 = tk.Entry(root)
        textbox2.insert(tk.END, str(param))  # 设置默认值
        textbox2.grid(row=i+1, column=3, padx=(10, 10), pady=(5, 0))  # 调整文本框的位置和间距
        controls.append((label2, textbox2))

        # 创建标签
        label3 = tk.Label(root, text=f'第{i}段回旋线类型:')
        label3.grid(row=i+1, column=4, padx=(10, 10), pady=(5, 0), sticky='e')  # 调整标签的位置和间距
        # 创建下拉框
        combobox = ttk.Combobox(root, values=["基本对称", "基本非对称", "S型", "卵型", "凸型", "C型", "复合型"])
        combobox.grid(row=i+1, column=5, padx=(10, 10), pady=(5, 0))  # 调整下拉框的位置和间距
        combobox.current(0)  # 设置默认选项
        controls.append((label3, combobox))
        # 绑定文本框和下拉框的事件处理函数
        textbox1.bind("<KeyRelease>", update)
        textbox2.bind("<KeyRelease>", update)
        combobox.bind("<<ComboboxSelected>>", update)
  

    #for slider in sliders:
    #    slider.on_changed(update)

    canvas.draw()
    root.mainloop()    
    #plt.show()


def zero_segments_calculator(x, f, zero_segments, A=10):
    x_segments = []
    y_segments = []
    # 根据 zero_segments 绘制不同颜色的样条曲线
    for segment in zero_segments:
        start_index, end_index = segment
        x_segment = x[start_index:end_index+1]
        y_segment = f(x_segment)
        x_segments.append(x_segment)
        y_segments.append(y_segment)
        #plt.plot(x_segment, y_segment, color='red', label=f'Zero derivative segment from {start_index} to {end_index}')
    return x_segments, y_segments

def bow_calculator(x, y, B=10):
    bows = []
    n = len(x)
    
    for i in range(0, n, B):
        if i > 0 and i < n - 1:  # 确保不越界
            # 获取当前点及其前后三个点的坐标
            x_points = x[max(i - 1, 0):min(i + 2, n)]
            y_points = y[max(i - 1, 0):min(i + 2, n)]
            
            # 计算连线斜率
            slope = (y_points[2] - y_points[0]) / (x_points[2] - x_points[0])
            
            # 计算连线的垂线斜率
            perpendicular_slope = -1 / slope if slope != 0 else np.inf
            
            # 计算垂线交点的坐标
            x_intersect = (x_points[1] + (y_points[1] - y_points[0]) / perpendicular_slope) / (1 + perpendicular_slope**2)
            y_intersect = y_points[1] + perpendicular_slope * (x_intersect - x_points[1])
            
            # 计算垂线长度
            bow_height = np.sqrt((x_points[1] - x_intersect)**2 + (y_points[1] - y_intersect)**2)
            bows.append(bow_height)
    
    return bows

def non_zero_segments_calculator(x, f, non_zero_segments, A, B):
    x_fits = []
    y_fits = []
    errors = []
    centers = []
    # 其他段的样条曲线颜色
    for i, segment in enumerate(non_zero_segments):
        start_index, end_index = segment
        x_segment = x[start_index:end_index+1]
        y_segment = f(x_segment)
        #plt.plot(x_segment, y_segment, color='blue', label=f'Non-Zero derivative segment from {start_index} to {end_index}')
        # 计算曲率半径
        curvature = calculate_curvature(f, x_segment)
        # 找到每个 segment 内的最大曲率及其对应的 x 值或索引
        if B is None:
            center = np.argmax(curvature)
        else:
            center = B[i]
        x_fit_circle, y_fit_circle, circle_error = circle_fit(center, x_segment, y_segment, A[i])  
        #plt.plot(x_fit_circle, y_fit_circle, 'g-', linewidth=3,  label=f'Fitted circle for segment {max_curvature_index}')
        #plt.axis('equal')  # 设置x轴和y轴比例相同，保证圆的形状正确显示
        # 标记每个曲率半径
        #for i in range(len(x_segment)):
        #     if i % 40 == 0:
        #        plt.scatter(x_segment[i], y_segment[i], color='green', s=10)
        #        plt.text(x_segment[i], y_segment[i], f'{curvature[i]:.2f}', fontsize=8, color='green')
        if x_fit_circle is not None and y_fit_circle is not None:
            x_fits.append(x_fit_circle)
            y_fits.append(y_fit_circle)
            errors.append(circle_error)
            centers.append(center)
    return x_fits, y_fits, errors, centers

def circle_fit(curvature_center, x, y, A):
    # 提取前后十个点的序号
    fit_start = max(0, curvature_center - A)
    fit_end = min(len(x), curvature_center + A)
    x_fit = x[fit_start:fit_end]
    y_fit = y[fit_start:fit_end]
    # 检查拟合点的数量
    if len(x_fit) < 3:  # 至少需要三个点来拟合圆
        print('拟合点数量不足')
        return None, None, None   
    # 拟合圆
    center_estimate = np.mean(x_fit), np.mean(y_fit)
    center, ier = leastsq(f_2, center_estimate, args=(x_fit, y_fit))
    xc, yc = center
    Ri = calc_R(xc, yc, x_fit, y_fit)
    R = Ri.mean()
    
    # 计算拟合误差
    fitting_error = np.sum((Ri - R)**2)
    
    # 绘制拟合圆的部分圆
    theta_start = np.arctan2(y_fit[0] - yc, x_fit[0] - xc)
    theta_end = np.arctan2(y_fit[-1] - yc, x_fit[-1] - xc)
    theta_fit = np.linspace(theta_start, theta_end, 100)
    x_fit_circle = xc + R * np.cos(theta_fit)
    y_fit_circle = yc + R * np.sin(theta_fit)
    return x_fit_circle, y_fit_circle, fitting_error

# 定义计算曲率的函数
def calculate_curvature(f, x):
    # 计算一阶和二阶导数
    f_prime = f.derivative()
    f_second = f_prime.derivative()
    y_prime_segment = f_prime(x)
    y_second_segment = f_second(x)
    # 计算曲率
    curvature = np.abs(y_second_segment) / (1 + y_prime_segment**2)**(3/2)
    return curvature

def compute_derivatives(f, x):
    f_prime = f.derivative()
    f_second = f_prime.derivative()
    f_third = f_second.derivative()
    y_prime = f_prime(x)
    y_second = f_second(x)
    y_third = f_third(x)  
    return y_prime, y_second, y_third


def find_zero_second_derivatives(x, y_dedrivative, tolerance):
    # 找到n阶导数为零的段落
    zero_crossings = np.where(np.abs(y_dedrivative) < tolerance)[0]
    zero_segments = np.split(zero_crossings, np.where(np.diff(zero_crossings) != 1)[0] + 1)
    zero_segments = [(segment[0], segment[-1]) for segment in zero_segments if len(segment) > 0]

    # 找到n阶导数不为零的段落
    nonzero_segments = []
    start_index = 0
    for segment in zero_segments:
        end_index = segment[0]
        if start_index < end_index:
            nonzero_segments.append((start_index, end_index))
        start_index = segment[-1]
    if start_index < len(x):
        nonzero_segments.append((start_index, len(x)))

    segments_to_remove_index = []  # 记录需要删除的段落索引
    for i, nonzero_segment in enumerate(nonzero_segments):
        if nonzero_segment[1] - nonzero_segment[0] > 2:
            continue
        # 找到zero_segments中尾数等于nonzero_segment[0]的数组，将它的尾数改为nonzero_segment[1] 
        segment_index = next((i for i, segment in enumerate(zero_segments) if segment[1] == nonzero_segment[0]), None)
        if segment_index is not None:
            # 将zero_segment的尾数改为nonzero_segment的尾数
            zero_segments[segment_index] = (zero_segments[segment_index][0], nonzero_segment[1])
            segments_to_remove_index.append(i)  # 记录需要删除的段落索引

    # 删除需要删除的段落
    for index in sorted(segments_to_remove_index, reverse=True):
        del nonzero_segments[index]

    return zero_segments, nonzero_segments


def main():
    result = extract_data_from_excel(file_path)
    normalized_data = normalize_data(result)
    folder_path = r'F:\0025 多源数据道路\RoadFinder\traindata'
    segments = segment_data(result)
    stake_values = result[:, 1].astype(float)
    lenght = stake_values[-1] - stake_values[0]
    x_values = result[:, 2].astype(float) 
    y_values = result[:, 3].astype(float)   
    Spline_fit_byXY(lenght, x_values, y_values, show=True)
    print(result)

if __name__ == "__main__":
    main()